"""
Module de mise à jour BOM avec gestion des statuts
Implémente la logique complète de comparaison et mise à jour selon les statuts D, X, 0, NaN
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import logging

class BOMStatusUpdater:
    """
    Classe principale pour la mise à jour des BOM avec gestion des statuts
    """
    
    def __init__(self, master_bom_path=None):
        """
        Initialise l'updater
        
        Args:
            master_bom_path (str): Chemin vers le fichier Master BOM
        """
        self.master_bom_path = master_bom_path or "master_bom_cleaned.xlsx"
        self.master_df = None
        self.cleaned_df = None
        self.report_data = []
        
        # Configuration des colonnes possibles
        self.possible_pn_names = [
            'Part Number', 'PN', 'Réf Composant', 'Reference', 
            'Ref', 'Part_Number', 'Référence', 'Part_Num'
        ]
        self.possible_project_names = [
            'Projet', 'Project', 'Prj', 'Programme', 'Program'
        ]
        self.possible_status_names = [
            'Statut', 'Status', 'État', 'State', 'Etat'
        ]
        
        # Configuration du logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    
    def load_master_bom(self):
        """
        Charge le Master BOM depuis le fichier
        
        Returns:
            bool: True si chargement réussi, False sinon
        """
        try:
            if not os.path.exists(self.master_bom_path):
                self.logger.error(f"Fichier Master BOM non trouvé: {self.master_bom_path}")
                return False
            
            self.master_df = pd.read_excel(self.master_bom_path)
            self.logger.info(f"Master BOM chargé: {len(self.master_df)} lignes")
            return True
            
        except Exception as e:
            self.logger.error(f"Erreur lors du chargement du Master BOM: {str(e)}")
            return False
    
    def load_cleaned_file(self, file_path):
        """
        Charge le fichier Excel nettoyé
        
        Args:
            file_path (str): Chemin vers le fichier nettoyé
            
        Returns:
            bool: True si chargement réussi, False sinon
        """
        try:
            self.cleaned_df = pd.read_excel(file_path)
            self.logger.info(f"Fichier nettoyé chargé: {len(self.cleaned_df)} lignes")
            return True
            
        except Exception as e:
            self.logger.error(f"Erreur lors du chargement du fichier nettoyé: {str(e)}")
            return False
    
    def _find_column(self, df, possible_names):
        """
        Trouve une colonne dans le DataFrame selon les noms possibles
        
        Args:
            df (DataFrame): DataFrame à analyser
            possible_names (list): Liste des noms possibles
            
        Returns:
            str: Nom de la colonne trouvée ou None
        """
        for col in df.columns:
            for name in possible_names:
                if name.lower() in col.lower():
                    return col
        return None
    
    def _xlookup_status(self, part_number, project=""):
        """
        Effectue un XLOOKUP pour trouver le statut dans le Master BOM
        
        Args:
            part_number (str): Numéro de pièce
            project (str): Nom du projet
            
        Returns:
            str: Statut trouvé ('D', 'X', '0', 'NaN')
        """
        if self.master_df is None:
            return "NaN"
        
        # Identifier les colonnes dans Master BOM
        master_pn_col = self._find_column(self.master_df, self.possible_pn_names)
        master_project_col = self._find_column(self.master_df, self.possible_project_names)
        master_status_col = self._find_column(self.master_df, self.possible_status_names)
        
        if not master_pn_col:
            return "NaN"
        
        # Recherche par Part Number
        mask = self.master_df[master_pn_col].astype(str).str.strip() == str(part_number).strip()
        
        # Affiner par projet si disponible
        if master_project_col and project:
            project_mask = self.master_df[master_project_col].astype(str).str.strip() == str(project).strip()
            mask = mask & project_mask
        
        matches = self.master_df[mask]
        
        if matches.empty:
            return "NaN"  # PN introuvable
        elif len(matches) > 1:
            return "0"    # Doublon/ambigu
        else:
            # Un seul match trouvé
            if master_status_col and master_status_col in matches.columns:
                status = matches[master_status_col].iloc[0]
                if pd.isna(status) or str(status).strip() == "":
                    return "NaN"
                return str(status).strip()
            else:
                return "NaN"
    
    def process_file(self):
        """
        Traite le fichier ligne par ligne selon la logique des statuts
        
        Returns:
            dict: Résultats du traitement
        """
        if self.cleaned_df is None or self.master_df is None:
            self.logger.error("Fichiers non chargés")
            return None
        
        # Identifier les colonnes dans le fichier nettoyé
        pn_col = self._find_column(self.cleaned_df, self.possible_pn_names)
        project_col = self._find_column(self.cleaned_df, self.possible_project_names)
        
        if not pn_col:
            self.logger.error("Colonne Part Number non trouvée dans le fichier nettoyé")
            return None
        
        # Ajouter colonne Notes si elle n'existe pas
        if 'Notes' not in self.cleaned_df.columns:
            self.cleaned_df['Notes'] = ""
        
        # Copie de travail
        working_df = self.cleaned_df.copy()
        
        # Statistiques
        stats = {'D': 0, 'X': 0, '0': 0, 'NaN': 0}
        
        # Traitement ligne par ligne
        for idx, row in working_df.iterrows():
            part_number = row[pn_col] if pd.notna(row[pn_col]) else ""
            project = row[project_col] if project_col and pd.notna(row[project_col]) else ""
            
            # XLOOKUP pour obtenir le statut
            status = self._xlookup_status(part_number, project)
            stats[status] += 1
            
            # Actions selon le statut
            if status == "D":
                # Ne rien faire
                continue
                
            elif status == "X":
                # Mettre à jour Master BOM et ajouter commentaire
                self._update_master_status(part_number, project, "D")
                working_df.at[idx, 'Notes'] = "Statut X remplacé par D"
                self._add_to_report(idx, part_number, project, "X", "Statut mis à jour de X vers D")
                
            elif status == "0":
                # Ajouter nouvelle ligne pour doublon
                new_row = self._create_new_row(part_number, project, "Doublon ou incertain - Vérification manuelle requise")
                working_df = pd.concat([working_df, pd.DataFrame([new_row])], ignore_index=True)
                self._add_to_report(len(working_df)-1, part_number, project, "0", "Nouvelle ligne ajoutée - doublon détecté")
                
            elif status == "NaN":
                # Gérer les cas NaN (erreur format ou PN introuvable)
                if part_number == "" or not self._is_valid_format(part_number):
                    # Erreur de format
                    new_row = self._create_new_row("", "", "Erreur de format – à corriger manuellement")
                    working_df = pd.concat([working_df, pd.DataFrame([new_row])], ignore_index=True)
                    self._add_to_report(len(working_df)-1, part_number, project, "NaN", "Erreur de format détectée")
                else:
                    # PN introuvable
                    new_row = self._create_new_row(part_number, project, "PN inconnu – insertion possible")
                    working_df = pd.concat([working_df, pd.DataFrame([new_row])], ignore_index=True)
                    self._add_to_report(len(working_df)-1, part_number, project, "NaN", "PN introuvable dans Master BOM")
        
        return {
            'updated_df': working_df,
            'stats': stats,
            'report_data': self.report_data
        }

    def _update_master_status(self, part_number, project, new_status):
        """
        Met à jour le statut dans le Master BOM

        Args:
            part_number (str): Numéro de pièce
            project (str): Projet
            new_status (str): Nouveau statut
        """
        master_pn_col = self._find_column(self.master_df, self.possible_pn_names)
        master_project_col = self._find_column(self.master_df, self.possible_project_names)
        master_status_col = self._find_column(self.master_df, self.possible_status_names)

        if not master_pn_col or not master_status_col:
            return

        # Recherche de la ligne à mettre à jour
        mask = self.master_df[master_pn_col].astype(str).str.strip() == str(part_number).strip()

        if master_project_col and project:
            project_mask = self.master_df[master_project_col].astype(str).str.strip() == str(project).strip()
            mask = mask & project_mask

        # Mise à jour
        if mask.any():
            self.master_df.loc[mask, master_status_col] = new_status

    def _create_new_row(self, part_number, project, comment):
        """
        Crée une nouvelle ligne avec les informations de base

        Args:
            part_number (str): Numéro de pièce
            project (str): Projet
            comment (str): Commentaire

        Returns:
            dict: Nouvelle ligne
        """
        # Identifier les colonnes dans le fichier nettoyé
        pn_col = self._find_column(self.cleaned_df, self.possible_pn_names)
        project_col = self._find_column(self.cleaned_df, self.possible_project_names)

        new_row = {}

        # Remplir toutes les colonnes avec des valeurs vides
        for col in self.cleaned_df.columns:
            new_row[col] = ""

        # Remplir les colonnes spécifiques
        if pn_col:
            new_row[pn_col] = part_number
        if project_col:
            new_row[project_col] = project

        # Laisser vides: Prix, Description, Fournisseur
        price_cols = ['Prix', 'Price', 'Coût', 'Cost']
        desc_cols = ['Description', 'Désignation', 'Desc']
        supplier_cols = ['Fournisseur', 'Supplier', 'Vendor']

        for col in self.cleaned_df.columns:
            if any(name.lower() in col.lower() for name in price_cols + desc_cols + supplier_cols):
                new_row[col] = ""

        # Ajouter le commentaire
        new_row['Notes'] = comment

        return new_row

    def _is_valid_format(self, part_number):
        """
        Vérifie si le format du Part Number est valide

        Args:
            part_number (str): Part Number à vérifier

        Returns:
            bool: True si valide, False sinon
        """
        if not part_number or pd.isna(part_number):
            return False

        part_str = str(part_number).strip()

        # Vérifications de base
        if len(part_str) < 2:
            return False

        # Vérifier qu'il contient au moins des caractères alphanumériques
        if not any(c.isalnum() for c in part_str):
            return False

        return True

    def _add_to_report(self, row_index, part_number, project, status, action):
        """
        Ajoute une entrée au rapport

        Args:
            row_index (int): Index de la ligne
            part_number (str): Part Number
            project (str): Projet
            status (str): Statut trouvé
            action (str): Action effectuée
        """
        self.report_data.append({
            'Ligne': row_index + 1,
            'Part Number': part_number,
            'Projet': project,
            'Statut': status,
            'Action': action,
            'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    def save_updated_file(self, updated_df, output_path=None):
        """
        Sauvegarde le fichier mis à jour avec formatage

        Args:
            updated_df (DataFrame): DataFrame mis à jour
            output_path (str): Chemin de sortie (optionnel)

        Returns:
            str: Chemin du fichier sauvegardé
        """
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"Update_{timestamp}.xlsx"

        try:
            # Sauvegarder avec openpyxl pour le formatage
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                updated_df.to_excel(writer, sheet_name='Updated_BOM', index=False)

                # Accéder au workbook pour le formatage
                workbook = writer.book
                worksheet = writer.sheets['Updated_BOM']

                # Appliquer le formatage
                self._apply_formatting(worksheet, updated_df)

            self.logger.info(f"Fichier sauvegardé: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Erreur lors de la sauvegarde: {str(e)}")
            return None

    def _apply_formatting(self, worksheet, df):
        """
        Applique le formatage au fichier Excel

        Args:
            worksheet: Feuille Excel
            df (DataFrame): DataFrame pour référence
        """
        # Couleurs pour les différents statuts
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        # Parcourir les lignes pour appliquer le formatage
        for row_idx, row in df.iterrows():
            excel_row = row_idx + 2  # +2 car Excel commence à 1 et on a les headers

            notes = str(row.get('Notes', ''))

            # Formatage selon le contenu des notes
            if 'Doublon' in notes or 'À vérifier' in notes:
                # Rouge pour les doublons
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=excel_row, column=col).fill = red_fill

            elif 'Erreur de format' in notes or 'PN inconnu' in notes:
                # Orange pour les erreurs
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=excel_row, column=col).fill = orange_fill

            elif 'Statut X remplacé' in notes:
                # Jaune pour les mises à jour
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=excel_row, column=col).fill = yellow_fill

    def save_master_bom(self, output_path=None):
        """
        Sauvegarde le Master BOM mis à jour

        Args:
            output_path (str): Chemin de sortie (optionnel)

        Returns:
            str: Chemin du fichier sauvegardé
        """
        if self.master_df is None:
            self.logger.error("Master BOM non chargé")
            return None

        if output_path is None:
            output_path = self.master_bom_path

        try:
            self.master_df.to_excel(output_path, index=False)
            self.logger.info(f"Master BOM sauvegardé: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Erreur lors de la sauvegarde du Master BOM: {str(e)}")
            return None

    def generate_report(self, output_format='csv', output_path=None):
        """
        Génère un rapport final des modifications

        Args:
            output_format (str): Format du rapport ('csv' ou 'pdf')
            output_path (str): Chemin de sortie (optionnel)

        Returns:
            str: Chemin du rapport généré
        """
        if not self.report_data:
            self.logger.warning("Aucune donnée de rapport disponible")
            return None

        # Créer DataFrame du rapport
        report_df = pd.DataFrame(self.report_data)

        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            extension = 'csv' if output_format == 'csv' else 'xlsx'
            output_path = f"Rapport_Modifications_{timestamp}.{extension}"

        try:
            if output_format.lower() == 'csv':
                report_df.to_csv(output_path, index=False, encoding='utf-8-sig')
            else:
                # Sauvegarder en Excel avec formatage
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    report_df.to_excel(writer, sheet_name='Rapport', index=False)

                    # Ajouter une feuille de statistiques
                    stats_data = []
                    for status, count in self._calculate_final_stats().items():
                        stats_data.append({'Statut': status, 'Nombre': count})

                    stats_df = pd.DataFrame(stats_data)
                    stats_df.to_excel(writer, sheet_name='Statistiques', index=False)

            self.logger.info(f"Rapport généré: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Erreur lors de la génération du rapport: {str(e)}")
            return None

    def _calculate_final_stats(self):
        """
        Calcule les statistiques finales

        Returns:
            dict: Statistiques par statut
        """
        stats = {'D': 0, 'X': 0, '0': 0, 'NaN': 0}

        for entry in self.report_data:
            status = entry.get('Statut', 'NaN')
            if status in stats:
                stats[status] += 1

        return stats

    def run_complete_process(self, cleaned_file_path, output_dir=None):
        """
        Exécute le processus complet de mise à jour

        Args:
            cleaned_file_path (str): Chemin vers le fichier nettoyé
            output_dir (str): Répertoire de sortie (optionnel)

        Returns:
            dict: Résultats complets du processus
        """
        if output_dir is None:
            output_dir = os.path.dirname(cleaned_file_path) or "."

        # 1. Charger les fichiers
        if not self.load_master_bom():
            return {'success': False, 'error': 'Impossible de charger le Master BOM'}

        if not self.load_cleaned_file(cleaned_file_path):
            return {'success': False, 'error': 'Impossible de charger le fichier nettoyé'}

        # 2. Traiter le fichier
        results = self.process_file()
        if not results:
            return {'success': False, 'error': 'Erreur lors du traitement'}

        # 3. Sauvegarder le fichier mis à jour
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        updated_file_path = os.path.join(output_dir, f"Update_{timestamp}.xlsx")
        saved_file = self.save_updated_file(results['updated_df'], updated_file_path)

        # 4. Sauvegarder le Master BOM mis à jour
        master_saved = self.save_master_bom()

        # 5. Générer le rapport
        report_path = os.path.join(output_dir, f"Rapport_Modifications_{timestamp}.xlsx")
        report_file = self.generate_report('xlsx', report_path)

        return {
            'success': True,
            'updated_file': saved_file,
            'master_bom_updated': master_saved,
            'report_file': report_file,
            'stats': results['stats'],
            'total_processed': len(self.cleaned_df),
            'total_modified': len(self.report_data)
        }
